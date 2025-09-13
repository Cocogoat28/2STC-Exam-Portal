from __future__ import annotations
from django.contrib import admin, messages
from django.db import transaction
from django.shortcuts import render, redirect
from django.urls import path, reverse
from django.http import HttpResponse, HttpResponseForbidden
from django.template.response import TemplateResponse
import time
from io import BytesIO

from django.utils import timezone

from .models import Candidate, Question, Answer
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# ------------ Excel helpers ------------

REQUIRED_COLS = {"army_no", "exam_type", "question", "answer"}
KNOWN_COLS = {
    "s_no", "name", "center", "photo", "fathers_name", "dob", "trade", "rank", "army_no", "adhaar_no",
    # NEW:
    "primary_qualification", "primary_duration", "primary_credits",
    "secondary_qualification", "secondary_duration", "secondary_credits",
    # existing:
    "nsqf_level", "training_center", "district", "state", "viva_1", "viva_2",
    "practical_1", "practical_2", "exam_type", "question", "answer",
    "correct_answer", "max_marks", "part",
}


def _normalize_header(val: str) -> str:
    # robust normalization: accept None and numeric values
    if val is None:
        return ""
    key = str(val).strip().lower().replace(".", "_").replace(" ", "_")
    mapping = {
        "s_no": "s_no", "sno": "s_no", "s_no.": "s_no", "s_number": "s_no",
        "fathers_name": "fathers_name", "father_name": "fathers_name",
        "army_no": "army_no", "army_number": "army_no",
        "adhaar_no": "adhaar_no", "aadhar_no": "adhaar_no",

        "primary_qualification": "primary_qualification",
        "primary qualification": "primary_qualification",
        "primary_duration": "primary_duration",
        "primary duration": "primary_duration",
        "primary_credits": "primary_credits",
        "primary credits": "primary_credits",

        "secondary_qualification": "secondary_qualification",
        "secondary qualification": "secondary_qualification",
        "secondary_duration": "secondary_duration",
        "secondary duration": "secondary_duration",
        "secondary_credits": "secondary_credits",
        "secondary credits": "secondary_credits",

        "nsqf_level": "nsqf_level", "nsqf": "nsqf_level", "nsqflevel": "nsqf_level",
        "training_center": "training_center", "centre_of_training": "training_center",

        # Excel variants
        "center": "center", "centre": "center",
        "trade": "trade", "trd": "trade",
    }
    return mapping.get(key, key)


def _read_rows_from_excel(file):
    """
    Robust header detection:
    - scans the top N rows to find a header row which contains the REQUIRED_COLS after normalization.
    - returns generator of dicts mapping normalized header -> cell value for each data row after the header.
    """
    wb = load_workbook(file, data_only=True)
    ws = wb.worksheets[0]

    # scan first few rows to find header row
    header_row_index = None
    max_header_scan = 10
    for r in range(1, min(max_header_scan, ws.max_row) + 1):
        # values_only=True returns plain values, so don't access .value
        row_values = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
        headers = [_normalize_header(v) for v in row_values]
        header_index = {h: idx for idx, h in enumerate(headers) if h}
        if REQUIRED_COLS.issubset(set(header_index.keys())):
            header_row_index = r
            break

    if header_row_index is None:
        # fallback: use first row as header (old behavior) but try to produce a useful error
        # here we want Cell objects so use values_only=False
        header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        headers = [_normalize_header(c.value) for c in header_cells]
        header_index = {h: idx for idx, h in enumerate(headers) if h}
        missing = REQUIRED_COLS - set(header_index)
        if missing:
            raise ValueError(f"Missing required columns in Excel (couldn't auto-detect header row). Missing: {', '.join(missing)}")
        header_row_index = 1
    else:
        # rebuild header_index for chosen header_row_index (we already got values_only row earlier, but re-read as values)
        header_cells = next(ws.iter_rows(min_row=header_row_index, max_row=header_row_index, values_only=True))
        headers = [_normalize_header(v) for v in header_cells]
        header_index = {h: idx for idx, h in enumerate(headers) if h}

    # iterate data rows starting after header_row_index
    for row in ws.iter_rows(min_row=header_row_index + 1, values_only=True):
        data = {}
        for key, idx in header_index.items():
            # guard: row may be shorter than header row
            try:
                data[key] = row[idx]
            except IndexError:
                data[key] = None
        yield data



def _get_or_create_question(exam_type, text, correct, max_marks, part=None):
    if part:
        part = str(part).strip().upper()
    q = Question.objects.filter(exam_type=exam_type, question=text).first()
    correct_clean = (correct or "")
    if isinstance(correct_clean, str) and correct_clean.strip().lower() == "null":
        correct_clean = None

    if q is None:
        q = Question.objects.create(
            exam_type=exam_type,
            question=text,
            part=part,
            correct_answer=correct_clean,
            max_marks=max_marks or 0,
        )
    else:
        q.correct_answer = correct_clean
        q.max_marks = max_marks or 0
        q.part = part or q.part
        q.save()
    return q


# ------------ Custom Admins ------------

class AnswerInline(admin.TabularInline):
    model = Answer
    extra = 0


# (Assume imports and helpers defined earlier remain the same)
@admin.register(Candidate)
class CandidateAdmin(admin.ModelAdmin):
    change_list_template = "admin/exams/candidate/change_list.html"
    change_form_template = "admin/exams/candidate/change_form.html"
    readonly_fields = ("viva_1", "viva_2", "practical_1", "practical_2")
    list_display = ("army_no", "name", "center", "trade", "total_primary", "total_secondary", "grand_total", "is_checked")
    list_filter = ("center", "trade", "is_checked")
    search_fields = ("army_no", "name", "rank", "fathers_name", "district", "state", "trade")

    # ✅ Add custom actions (only one 3-in-1)
    actions = [
        "export_selected_results",          # 3-in-1 workbook (PRIMARY, SECONDARY, COMBINED)
        "export_selected_evaluation_list",  # single-sheet checked metadata
        "export_selected_export_all",       # single-sheet detailed EXPORT_ALL
    ]

    def get_urls(self):
        urls = super().get_urls()
        custom = [
            path(
                "export-evaluation/",
                self.admin_site.admin_view(self.export_evaluation_page),
                name="exams_export_evaluation_page",
            ),
            path(
                "import-excel/",
                self.admin_site.admin_view(self.import_excel_view),
                name="exams_candidate_import_excel",
            ),
            path(
                "export-results-excel/",
                self.admin_site.admin_view(self.export_results_excel_view),
                name="exams_export_results_excel",
            ),
            path(
                "export-evaluation-sheet/",
                self.admin_site.admin_view(self.export_evaluation_sheet_view),
                name="exams_export_evaluation_sheet",
            ),
            path(
                "export-all-sheet/",
                self.admin_site.admin_view(self.export_all_sheet_view),
                name="exams_export_all_sheet",
            ),
            path(
                "<int:candidate_id>/save-grades/",
                self.admin_site.admin_view(self.save_grades_view),
                name="exams_candidate_save_grades",
            ),
            path(
                "<int:candidate_id>/grade-answers/",
                self.admin_site.admin_view(self.grade_answers_view),
                name="exams_candidate_grade_answers",
            ),
        ]
        return custom + urls

    # ---------- Candidate change form ----------
    def change_view(self, request, object_id, form_url="", extra_context=None):
        cand = Candidate.objects.get(pk=object_id)
        answers = Answer.objects.filter(candidate=cand).select_related("question")

        primary = [a for a in answers if a.question.exam_type.lower() == "primary"]
        secondary = [a for a in answers if a.question.exam_type.lower() == "secondary"]

        extra_context = extra_context or {}
        extra_context["primary_answers"] = primary
        extra_context["secondary_answers"] = secondary
        extra_context["viva_total"] = cand.viva_1 + cand.viva_2
        extra_context["practical_total"] = cand.practical_1 + cand.practical_2
        extra_context["show_grade_button"] = True

        return super().change_view(request, object_id, form_url, extra_context=extra_context)

    # ---------- Grade Answers View ----------
    def grade_answers_view(self, request, candidate_id):
        if not request.user.has_perm('exams.change_answer'):
            return HttpResponseForbidden("You don't have permission to grade answers")

        cand = Candidate.objects.get(pk=candidate_id)
        answers = Answer.objects.filter(candidate=cand).select_related("question")

        # Auto-marking logic
        # Auto-marking logic — only for MCQ (A,B) and True/False (F)
        for ans in answers:
            part = (ans.question.part or '').strip().upper()
            if part not in ('A', 'B', 'F'):
                # do not auto-mark free-text answers (C, D, E)
                continue

            cand_ans = (ans.answer or "").strip().lower()
            corr_raw = (ans.question.correct_answer or "").strip().lower()
            if cand_ans and corr_raw:
                correct_list = [c.strip() for c in corr_raw.split(",")]
                if cand_ans in correct_list:
                    if ans.marks_obt is None or ans.marks_obt == 0:
                        ans.marks_obt = ans.question.max_marks
                        ans.save()

        primary_answers = [a for a in answers if a.question.exam_type.lower() == "primary"]
        secondary_answers = [a for a in answers if a.question.exam_type.lower() == "secondary"]

        def group_answers(ans_list):
            def norm(p): return (p or "").strip().upper()
            return {
                "MCQ": [a for a in ans_list if norm(a.question.part) in ("A", "B")],
                "Short Answer": [a for a in ans_list if norm(a.question.part) == "C"],
                "Fill in Blanks": [a for a in ans_list if norm(a.question.part) == "D"],
                "True/False": [a for a in ans_list if norm(a.question.part) == "F"],
                "Long Answer": [a for a in ans_list if norm(a.question.part) == "E"],
            }

        # ---------------------------
        # Block POST edits if grades already locked
        # ---------------------------
        if request.method == "POST":
            if cand.is_checked:
                # Candidate already checked: refuse modifications
                self.message_user(request, "Grades are locked for this candidate — no further edits allowed.", level=messages.WARNING)
                return redirect('admin:exams_candidate_changelist')

            for answer in answers:
                field_name = f"marks_{answer.id}"
                if field_name in request.POST:
                    try:
                        marks_value = request.POST[field_name].strip()
                        # Server-side guard: if no answer for D/E, disallow marks
                        ans_part = (answer.question.part or '').strip().upper()
                        ans_text = (answer.answer or '').strip()
                        # Prevent awarding marks for blank free-text parts (Short answer, Fill in blanks, Long answer)
                        if ans_part in ('C', 'D', 'E') and ans_text == '':
                            # ignore any provided marks and clear marks
                            answer.marks_obt = None
                            answer.save()
                            continue
                        if marks_value == "":
                            answer.marks_obt = None
                        else:
                            new_marks = int(marks_value)
                            if 0 <= new_marks <= answer.question.max_marks:
                                answer.marks_obt = new_marks
                        answer.save()
                    except ValueError:
                        pass

            # record checked metadata here as well
            cand.is_checked = True
            cand.checked_at = timezone.now()
            if request.user and request.user.is_authenticated:
                cand.checked_by = request.user
            cand.save()

            self.message_user(request, "Grades updated successfully", level=messages.SUCCESS)
            return redirect('admin:exams_candidate_changelist')

        primary_total_obtained = sum(a.marks_obt or 0 for a in primary_answers)
        secondary_total_obtained = sum(a.marks_obt or 0 for a in secondary_answers)

        all_marks_assigned = all(
            answer.marks_obt is not None and answer.marks_obt != 0
            for answer in answers
        )

        context = {
            **self.admin_site.each_context(request),
            "title": f"Grade Answers - {cand.name} ({cand.army_no}) - {cand.trade}",
            "candidate": cand,
            "primary_answers": primary_answers,
            "secondary_answers": secondary_answers,
            "primary_groups": group_answers(primary_answers),
            "secondary_groups": group_answers(secondary_answers),
            "primary_total_obtained": primary_total_obtained,
            "secondary_total_obtained": secondary_total_obtained,
            "all_marks_assigned": all_marks_assigned,
            "opts": self.model._meta,
            # new flag for template (read-only/view-only)
            "is_locked": cand.is_checked,
        }
        return TemplateResponse(request, "admin/exams/candidate/grade_answers.html", context)

    # ---------- Save Grades View ----------
    def save_grades_view(self, request, candidate_id):
        cand = Candidate.objects.get(pk=candidate_id)
        # If already locked, refuse to make any changes here too
        if cand.is_checked:
            self.message_user(request, "Grades are locked for this candidate — no further edits allowed.", level=messages.WARNING)
            return redirect('admin:exams_candidate_changelist')

        if request.method == "POST":
            for ans in Answer.objects.filter(candidate=cand):
                field_name = f"marks_{ans.id}"
                if field_name in request.POST:
                    try:
                        new_marks = int(request.POST[field_name])
                        ans.marks_obt = new_marks
                        ans.save()
                    except ValueError:
                        pass
            # record checked metadata
            cand.is_checked = True
            cand.checked_at = timezone.now()
            if request.user and request.user.is_authenticated:
                cand.checked_by = request.user
            cand.save()
            self.message_user(request, "Grades updated", level=messages.SUCCESS)
        return redirect('admin:exams_candidate_changelist')

    # ... the rest of your class (import/export helpers and _generate_excel etc.) remains unchanged ...

    # ---------- Import Excel ----------
    def import_excel_view(self, request):
        if request.method == "POST" and request.FILES.get("excel"):
            excel_file = request.FILES["excel"]
            created_candidates = updated_candidates = 0
            created_answers = updated_answers = 0
            created_questions = 0

            # helper to safely coerce values to strings for string fields
            def safe_str(v):
                if v is None:
                    return ""
                if isinstance(v, str):
                    return v.strip()
                return str(v).strip()

            try:
                with transaction.atomic():
                    seen_questions_before = set(Question.objects.values_list("id", flat=True))
                    for row in _read_rows_from_excel(excel_file):
                        army = safe_str(row.get("army_no"))
                        if not army:
                            continue

                        cand_defaults = {
                            "s_no": row.get("s_no") or 0,
                            "name": safe_str(row.get("name")),
                            "center": safe_str(row.get("center")),
                            "photo": row.get("photo") or None,
                            "fathers_name": safe_str(row.get("fathers_name")),
                            "dob": row.get("dob") or None,
                            "rank": safe_str(row.get("rank")),
                            "trade": safe_str(row.get("trade")).upper(),
                            "adhaar_no": safe_str(row.get("adhaar_no")),
                            "primary_qualification": safe_str(row.get("primary_qualification")),
                            "primary_duration": row.get("primary_duration") or 0,
                            "primary_credits": row.get("primary_credits") or 0,
                            "secondary_qualification": safe_str(row.get("secondary_qualification")),
                            "secondary_duration": row.get("secondary_duration") or 0,
                            "secondary_credits": row.get("secondary_credits") or 0,
                            "nsqf_level": row.get("nsqf_level") or 0,
                            "training_center": safe_str(row.get("training_center")),
                            "district": safe_str(row.get("district")),
                            "state": safe_str(row.get("state")),
                            "viva_1": row.get("viva_1") or 0,
                            "viva_2": row.get("viva_2") or 0,
                            "practical_1": row.get("practical_1") or 0,
                            "practical_2": row.get("practical_2") or 0,
                        }

                        cand, created = Candidate.objects.get_or_create(
                            army_no=army, defaults=cand_defaults
                        )
                        if not created:
                            for k, v in cand_defaults.items():
                                # update only if provided and different
                                if v not in (None, "") and getattr(cand, k) != v:
                                    setattr(cand, k, v)
                            cand.save()
                            updated_candidates += 1
                        else:
                            created_candidates += 1

                        q = _get_or_create_question(
                            exam_type=safe_str(row.get("exam_type")),
                            text=safe_str(row.get("question")),
                            correct=row.get("correct_answer"),
                            max_marks=row.get("max_marks") or 0,
                            part=row.get("part") or None,
                        )
                        if q.id not in seen_questions_before:
                            created_questions += 1
                            seen_questions_before.add(q.id)

                        ans_text = safe_str(row.get("answer"))
                        marks_raw = row.get("marks_obt")
                        # defensive int conversion (empty or None => 0)
                        try:
                            marks = int(marks_raw) if marks_raw not in (None, "") else 0
                        except (ValueError, TypeError):
                            marks = 0

                        # ----------------------------
                        # IMPORTANT FIX: create answer INSIDE the loop for each row
                        # ----------------------------
                        # Always create a new Answer row to preserve duplicates from Excel.
                        # Ensure your Answer model does not have a unique constraint on (candidate, question)
                        ans = Answer.objects.create(candidate=cand, question=q, answer=ans_text, marks_obt=int(marks))
                        created_answers += 1

                self.message_user(
                    request,
                    (
                        f"Import complete. "
                        f"Candidates: +{created_candidates} / updated {updated_candidates}. "
                        f"Questions: +{created_questions}. "
                        f"Answers: +{created_answers} / updated {updated_answers}."
                    ),
                    level=messages.SUCCESS,
                )
                return redirect("admin:exams_candidate_changelist")

            except Exception as e:
                self.message_user(request, f"Import failed: {e}", level=messages.ERROR)

        ctx = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Import candidates & answers from Excel",
        }
        return render(request, "admin/exams/candidate/import_excel.html", ctx)

    # ---------- Export ALL (button) ----------
    def export_results_excel_view(self, request):
        """
        Export ALL candidates (ignores filters). This uses _generate_excel which now
        produces only PRIMARY, SECONDARY and COMBINED sheets (no EXPORT_ALL included).
        """
        queryset = Candidate.objects.all()
        return self._generate_excel(queryset)

    # ... rest of your class unchanged ...
    # (I intentionally left the remaining methods exactly as you had them)


    # ---------- NEW action: Export selected queryset as 3-in-1 ----------
    def export_selected_results(self, request, queryset):
        """
        Admin action: export the selected/filtered queryset as the 3-in-1 workbook
        (PRIMARY MARKS STATEMENT, SECONDARY MARKS STATEMENT, COMBINED RESULTS).
        """
        return self._generate_excel(queryset)

    export_selected_results.short_description = "Export Final Result"

    # ---------- NEW action: Export selected queryset as Evaluation List (single sheet) ----------
    def export_selected_evaluation_list(self, request, queryset):
        """
        Admin action: export selected/filtered queryset as single-sheet evaluation list.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "EVALUATION_LIST"

        headers = [
            "S No", "Army No", "Name", "Centre", "Trade",
            "Primary Total", "Secondary Total", "Grand Total",
            "Is Checked", "Checked By", "Checked At"
        ]
        ws.append(headers)

        for idx, cand in enumerate(queryset.order_by("center", "army_no"), start=1):
            primary_total = cand.total_primary()
            secondary_total = cand.total_secondary()
            grand = cand.grand_total()
            is_checked = "Yes" if cand.is_checked else "No"
            checked_by = ""
            try:
                checked_by = cand.checked_by.get_username() if cand.checked_by else ""
            except Exception:
                checked_by = str(cand.checked_by) if cand.checked_by else ""

            checked_at = ""
            if getattr(cand, "checked_at", None):
                try:
                    checked_local = timezone.localtime(cand.checked_at)
                    checked_at = checked_local.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    checked_at = str(cand.checked_at)

            ws.append([
                idx, cand.army_no or "", cand.name or "", cand.center or "", cand.trade or "",
                primary_total, secondary_total, grand,
                is_checked, checked_by, checked_at
            ])

        # add basic formatting/borders
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            if cell.value:
                cell.font = bold_font
                cell.alignment = center_aligned
                cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="evaluation_list_selected.xlsx"'
        return response

    export_selected_evaluation_list.short_description = "Export Evaluation List (checked metadata)"

    # ---------- NEW action: Export selected queryset as EXPORT_ALL (single sheet) ----------
    def export_selected_export_all(self, request, queryset):
        """
        Admin action: export selected/filtered queryset as the EXPORT_ALL single-sheet.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "EXPORT_ALL"

        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = [
            "S No", "Army No", "Name", "Center", "Photo", "Father's Name", "DOB", "Rank", "Trade", "Aadhar Number",
            "Primary Qualification", "Primary Duration", "Primary Credits",
            "Secondary Qualification", "Secondary Duration", "Secondary Credits",
            "NSQF Level", "Training Centre", "District", "State",
            "Exam Type", "Question Part", "Question Text",
            "Correct Answer", "Max Marks", "Candidate Answer", "Marks Awarded",
            "Primary Total", "Secondary Total", "Grand Total",
            "Viva 1", "Viva 2", "Practical 1", "Practical 2",
            "Checked By", "Checked At"
        ]
        ws.append(headers)
        for cell in ws[1]:
            if cell.value:
                cell.font = bold_font
                cell.alignment = center_aligned
                cell.border = thin_border

        for idx, cand in enumerate(queryset.order_by("center", "army_no"), start=1):
            # compute totals
            primary_theory = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="primary")
            )
            primary_practical = cand.practical_1 or 0
            primary_viva = cand.viva_1 or 0
            primary_total = primary_theory + primary_practical + primary_viva

            secondary_theory = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="secondary")
            )
            secondary_practical = cand.practical_2 or 0
            secondary_viva = cand.viva_2 or 0
            secondary_total = secondary_theory + secondary_practical + secondary_viva

            grand_total = (primary_total or 0) + (secondary_total or 0)

            answers_qs = cand.answer_set.select_related("question").all()
            if not answers_qs:
                checked_by = ""
                try:
                    checked_by = cand.checked_by.get_username() if cand.checked_by else ""
                except Exception:
                    checked_by = str(cand.checked_by) if cand.checked_by else ""

                checked_at = ""
                if getattr(cand, "checked_at", None):
                    try:
                        checked_local = timezone.localtime(cand.checked_at)
                        checked_at = checked_local.strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        checked_at = str(cand.checked_at)

                ws.append([
                    idx,
                    cand.army_no or "", cand.name or "", cand.center or "", cand.photo or "",
                    cand.fathers_name or "", cand.dob or "", cand.rank or "", cand.trade or "", cand.adhaar_no or "",
                    cand.primary_qualification or "", cand.primary_duration or "", cand.primary_credits or "",
                    cand.secondary_qualification or "", cand.secondary_duration or "", cand.secondary_credits or "",
                    cand.nsqf_level or "", cand.training_center or "", cand.district or "", cand.state or "",
                    "", "", "", "", "", "", primary_total, secondary_total, grand_total,
                    cand.viva_1 or 0, cand.viva_2 or 0, cand.practical_1 or 0, cand.practical_2 or 0,
                    checked_by, checked_at
                ])
            else:
                for ans in answers_qs:
                    q = getattr(ans, "question", None)
                    question_text = (getattr(q, "question", "") or "")[:32767]
                    question_part = getattr(q, "part", "") or ""
                    question_exam_type = getattr(q, "exam_type", "") or ""
                    correct_answer = getattr(q, "correct_answer", "") or ""
                    max_marks = getattr(q, "max_marks", "") or ""

                    candidate_answer = getattr(ans, "answer", "") or ""
                    marks_awarded = getattr(ans, "marks_obt", "") or ""

                    checked_by = ""
                    try:
                        checked_by = cand.checked_by.get_username() if cand.checked_by else ""
                    except Exception:
                        checked_by = str(cand.checked_by) if cand.checked_by else ""

                    checked_at = ""
                    if getattr(cand, "checked_at", None):
                        try:
                            checked_local = timezone.localtime(cand.checked_at)
                            checked_at = checked_local.strftime("%Y-%m-%d %H:%M:%S")
                        except Exception:
                            checked_at = str(cand.checked_at)

                    ws.append([
                        idx,
                        cand.army_no or "", cand.name or "", cand.center or "", cand.photo or "",
                        cand.fathers_name or "", cand.dob or "", cand.rank or "", cand.trade or "", cand.adhaar_no or "",
                        cand.primary_qualification or "", cand.primary_duration or "", cand.primary_credits or "",
                        cand.secondary_qualification or "", cand.secondary_duration or "", cand.secondary_credits or "",
                        cand.nsqf_level or "", cand.training_center or "", cand.district or "", cand.state or "",
                        question_exam_type, question_part, question_text,
                        correct_answer, max_marks, candidate_answer, marks_awarded,
                        primary_total, secondary_total, grand_total,
                        cand.viva_1 or 0, cand.viva_2 or 0, cand.practical_1 or 0, cand.practical_2 or 0,
                        checked_by, checked_at
                    ])

        # add thin borders
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="export_all_selected.xlsx"'
        return response

    export_selected_export_all.short_description = "Export Answer Sheet"

    # ---------- Export page (button only) ----------
    def export_evaluation_page(self, request):
        ctx = {
            **self.admin_site.each_context(request),
            "opts": self.model._meta,
            "title": "Export Evaluation Data",
        }
        return render(request, "admin/exams/export_evaluation.html", ctx)

    # ---------- NEW: single-sheet export view (global) ----------
    def export_evaluation_sheet_view(self, request):
        queryset = Candidate.objects.all().order_by('center', 'army_no')

        wb = Workbook()
        ws = wb.active
        ws.title = "EVALUATION_LIST"

        headers = [
            "S No", "Army No", "Name", "Centre", "Trade",
            "Primary Total", "Secondary Total", "Grand Total",
            "Is Checked", "Checked By", "Checked At"
        ]
        ws.append(headers)

        for idx, cand in enumerate(queryset, start=1):
            primary_total = cand.total_primary()
            secondary_total = cand.total_secondary()
            grand = cand.grand_total()
            is_checked = "Yes" if cand.is_checked else "No"
            checked_by = ""
            try:
                checked_by = cand.checked_by.get_username() if cand.checked_by else ""
            except Exception:
                checked_by = str(cand.checked_by) if cand.checked_by else ""

            checked_at = ""
            if getattr(cand, "checked_at", None):
                try:
                    checked_local = timezone.localtime(cand.checked_at)
                    checked_at = checked_local.strftime("%Y-%m-%d %H:%M:%S")
                except Exception:
                    checked_at = str(cand.checked_at)

            ws.append([
                idx, cand.army_no or "", cand.name or "", cand.center or "", cand.trade or "",
                primary_total, secondary_total, grand,
                is_checked, checked_by, checked_at
            ])

        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            if cell.value:
                cell.font = bold_font
                cell.alignment = center_aligned
                cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="evaluation_list.xlsx"'
        return response

    # ---------- NEW: export only EXPORT_ALL as single sheet (global) ----------
    def export_all_sheet_view(self, request):
        queryset = Candidate.objects.all().order_by("center", "army_no")

        wb = Workbook()
        ws = wb.active
        ws.title = "EXPORT_ALL"

        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = [
            "S No", "Army No", "Name", "Center", "Photo", "Father's Name", "DOB", "Rank", "Trade", "Aadhar Number",
            "Primary Qualification", "Primary Duration", "Primary Credits",
            "Secondary Qualification", "Secondary Duration", "Secondary Credits",
            "NSQF Level", "Training Centre", "District", "State",
            "Exam Type", "Question Part", "Question Text",
            "Correct Answer", "Max Marks", "Candidate Answer", "Marks Awarded",
            "Primary Total", "Secondary Total", "Grand Total",
            "Viva 1", "Viva 2", "Practical 1", "Practical 2",
            "Checked By", "Checked At"
        ]
        ws.append(headers)
        for cell in ws[1]:
            if cell.value:
                cell.font = bold_font
                cell.alignment = center_aligned
                cell.border = thin_border

        for idx, cand in enumerate(queryset, start=1):
            # compute totals
            primary_theory = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="primary")
            )
            primary_practical = cand.practical_1 or 0
            primary_viva = cand.viva_1 or 0
            primary_total = primary_theory + primary_practical + primary_viva

            secondary_theory = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="secondary")
            )
            secondary_practical = cand.practical_2 or 0
            secondary_viva = cand.viva_2 or 0
            secondary_total = secondary_theory + secondary_practical + secondary_viva

            grand_total = (primary_total or 0) + (secondary_total or 0)

            answers_qs = cand.answer_set.select_related("question").all()
            if not answers_qs:
                checked_by = ""
                try:
                    checked_by = cand.checked_by.get_username() if cand.checked_by else ""
                except Exception:
                    checked_by = str(cand.checked_by) if cand.checked_by else ""

                checked_at = ""
                if getattr(cand, "checked_at", None):
                    try:
                        checked_local = timezone.localtime(cand.checked_at)
                        checked_at = checked_local.strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        checked_at = str(cand.checked_at)

                ws.append([
                    idx,
                    cand.army_no or "", cand.name or "", cand.center or "", cand.photo or "",
                    cand.fathers_name or "", cand.dob or "", cand.rank or "", cand.trade or "", cand.adhaar_no or "",
                    cand.primary_qualification or "", cand.primary_duration or "", cand.primary_credits or "",
                    cand.secondary_qualification or "", cand.secondary_duration or "", cand.secondary_credits or "",
                    cand.nsqf_level or "", cand.training_center or "", cand.district or "", cand.state or "",
                    "", "", "", "", "", "", primary_total, secondary_total, grand_total,
                    cand.viva_1 or 0, cand.viva_2 or 0, cand.practical_1 or 0, cand.practical_2 or 0,
                    checked_by, checked_at
                ])
            else:
                for ans in answers_qs:
                    q = getattr(ans, "question", None)
                    question_text = (getattr(q, "question", "") or "")[:32767]
                    question_part = getattr(q, "part", "") or ""
                    question_exam_type = getattr(q, "exam_type", "") or ""
                    correct_answer = getattr(q, "correct_answer", "") or ""
                    max_marks = getattr(q, "max_marks", "") or ""

                    candidate_answer = getattr(ans, "answer", "") or ""
                    marks_awarded = getattr(ans, "marks_obt", "") or ""

                    checked_by = ""
                    try:
                        checked_by = cand.checked_by.get_username() if cand.checked_by else ""
                    except Exception:
                        checked_by = str(cand.checked_by) if cand.checked_by else ""

                    checked_at = ""
                    if getattr(cand, "checked_at", None):
                        try:
                            checked_local = timezone.localtime(cand.checked_at)
                            checked_at = checked_local.strftime("%Y-%m-%d %H:%M:%S")
                        except Exception:
                            checked_at = str(cand.checked_at)

                    ws.append([
                        idx,
                        cand.army_no or "", cand.name or "", cand.center or "", cand.photo or "",
                        cand.fathers_name or "", cand.dob or "", cand.rank or "", cand.trade or "", cand.adhaar_no or "",
                        cand.primary_qualification or "", cand.primary_duration or "", cand.primary_credits or "",
                        cand.secondary_qualification or "", cand.secondary_duration or "", cand.secondary_credits or "",
                        cand.nsqf_level or "", cand.training_center or "", cand.district or "", cand.state or "",
                        question_exam_type, question_part, question_text,
                        correct_answer, max_marks, candidate_answer, marks_awarded,
                        primary_total, secondary_total, grand_total,
                        cand.viva_1 or 0, cand.viva_2 or 0, cand.practical_1 or 0, cand.practical_2 or 0,
                        checked_by, checked_at
                    ])

        # add thin borders
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    cell.border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="export_all.xlsx"'
        return response

    # ---------- Helper: Generate Excel (INSIDE the class) ----------
    def _generate_excel(self, queryset):
        """
        Produces a workbook with three sheets in this order:
        1) PRIMARY MARKS STATEMENT
        2) SECONDARY MARKS STATEMENT
        3) COMBINED RESULTS
        (This intentionally does NOT add the EXPORT_ALL sheet)
        """
        wb = Workbook()

        ws_primary = wb.active
        ws_primary.title = "PRIMARY MARKS STATEMENT"
        ws_secondary = wb.create_sheet(title="SECONDARY MARKS STATEMENT")
        ws_combined = wb.create_sheet(title="COMBINED RESULTS")

        bold_font = Font(bold=True)
        center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        # ----- Combined Sheet Formatting -----
        ws_combined.merge_cells("A1:A2")
        ws_combined.merge_cells("B1:B2")
        ws_combined.merge_cells("C1:C2")
        ws_combined.merge_cells("D1:D2")
        ws_combined.merge_cells("E1:E2")
        ws_combined.merge_cells("F1:F2")
        ws_combined.merge_cells("G1:K1")
        ws_combined.merge_cells("L1:P1")

        ws_combined["A1"] = "S No"
        ws_combined["B1"] = "Centre"
        ws_combined["C1"] = "Army No"
        ws_combined["D1"] = "Rk"
        ws_combined["E1"] = "Tde"
        ws_combined["F1"] = "Name"
        ws_combined["G1"] = "Primary-1"
        ws_combined["L1"] = "Secondary-1"

        sub_headers = [
            "Theory*", "Practical*", "Viva*", "Total", "Percentage (%)",
            "Theory*", "Practical*", "Viva*", "Total", "Percentage (%)"
        ]
        for i, val in enumerate(sub_headers, start=7):
            ws_combined.cell(row=2, column=i, value=val)

        for row in ws_combined.iter_rows(min_row=1, max_row=2):
            for cell in row:
                if cell.value:
                    cell.font = bold_font
                    cell.alignment = center_aligned
                    cell.border = thin_border

        # ----- Headers for Primary & Secondary -----
        primary_headers = [
            "S No", "Name of Candidate", "Photograph", "Father's Name", "Trade", "DOB",
            "Enrolment No", "Aadhar Number",
            "Primary Qualification", "Primary Duration", "Primary Credits",
            "NSQF Level", "Training Centre", "District", "State", "Percentage"
        ]
        secondary_headers = [
            "S No", "Name of Candidate", "Photograph", "Father's Name", "Trade", "DOB",
            "Enrolment No", "Aadhar Number",
            "Secondary Qualification", "Secondary Duration", "Secondary Credits",
            "NSQF Level", "Training Centre", "District", "State", "Percentage"
        ]

        ws_primary.append(primary_headers)
        ws_secondary.append(secondary_headers)

        for row in ws_primary.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = bold_font
                cell.alignment = center_aligned
                cell.border = thin_border

        for row in ws_secondary.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.font = bold_font
                cell.alignment = center_aligned
                cell.border = thin_border

        # ----- Fill Candidate Data -----
        for idx, cand in enumerate(queryset, start=1):
            primary_theory = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="primary")
            )
            primary_practical = cand.practical_1 or 0
            primary_viva = cand.viva_1 or 0
            primary_total = primary_theory + primary_practical + primary_viva
            primary_percentage = primary_total

            secondary_theory = sum(
                a.marks_obt or 0 for a in cand.answer_set.filter(question__exam_type__iexact="secondary")
            )
            secondary_practical = cand.practical_2 or 0
            secondary_viva = cand.viva_2 or 0
            secondary_total = secondary_theory + secondary_practical + secondary_viva
            secondary_percentage = secondary_total

            ws_combined.append([
                idx, cand.center or "", cand.army_no or "", cand.rank or "", cand.trade or "", cand.name or "",
                primary_theory, primary_practical, primary_viva, primary_total, primary_percentage,
                secondary_theory, secondary_practical, secondary_viva, secondary_total, secondary_percentage
            ])

            ws_primary.append([
                idx, cand.name or "", cand.photo or "", cand.fathers_name or "",
                cand.trade or "", cand.dob or "", cand.army_no or "", cand.adhaar_no or "",
                cand.primary_qualification or "", cand.primary_duration or "",
                cand.primary_credits or "", cand.nsqf_level or "", cand.training_center or "",
                cand.district or "", cand.state or "", primary_percentage
            ])

            ws_secondary.append([
                idx, cand.name or "", cand.photo or "", cand.fathers_name or "",
                cand.trade or "", cand.dob or "", cand.army_no or "", cand.adhaar_no or "",
                cand.secondary_qualification or "", cand.secondary_duration or "",
                cand.secondary_credits or "", cand.nsqf_level or "", cand.training_center or "",
                cand.district or "", cand.state or "", secondary_percentage
            ])

        # Add borders
        for ws in [ws_combined, ws_primary, ws_secondary]:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border

        # ✅ Return Excel file (3 sheets only)
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="results_3in1.xlsx"'
        return response
